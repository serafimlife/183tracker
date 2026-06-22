"""CSV/XLSX import — parse, validate, persist travel history in bulk."""

import csv
import html
import io
from dataclasses import dataclass, field
from datetime import date, datetime
from itertools import chain
from typing import Any, Iterable

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.repositories.stay_repository import StayRepository
from app.residency_engine import StayRecord, find_overlapping_stay, stay_duration_days
from app.residency_engine.duplicates import find_duplicate_entry
from app.services.residency_service import stays_to_records
from app.utils.countries import resolve_country


@dataclass
class ImportError:
    row: int
    message: str


@dataclass
class ImportResult:
    added_entries: int = 0
    countries: set[str] = field(default_factory=set)
    total_days: int = 0
    errors: list[ImportError] = field(default_factory=list)

    @property
    def message(self) -> str:
        lines = ["Import completed.", ""]
        lines.append(f"{self.added_entries} entries added")
        lines.append(f"{len(self.countries)} countries")
        lines.append(f"{self.total_days} total days imported")

        if self.errors:
            lines.append("")
            lines.append("Skipped rows:")
            for err in self.errors:
                lines.append(f"- Row {err.row}: {err.message}")

        return "\n".join(lines)


MAX_IMPORT_ROWS = 10_000


class ImportService:
    def __init__(self, session: AsyncSession) -> None:
        self._repo = StayRepository(session)

    async def import_csv(
        self,
        telegram_id: int,
        csv_text: str,
        *,
        date_format: str = "%Y-%m-%d",
    ) -> ImportResult:
        rows = csv.reader(io.StringIO(csv_text))
        return await self._import_rows(telegram_id, rows, date_format)

    async def import_xlsx(
        self,
        telegram_id: int,
        xlsx_content: bytes,
        *,
        date_format: str = "%Y-%m-%d",
    ) -> ImportResult:
        try:
            workbook = load_workbook(
                io.BytesIO(xlsx_content),
                read_only=True,
                data_only=True,
            )
            rows = list(workbook.active.iter_rows(values_only=True))
            workbook.close()
        except Exception:
            result = ImportResult()
            result.errors.append(ImportError(row=1, message="invalid XLSX file."))
            return result

        return await self._import_rows(telegram_id, rows, date_format)

    async def _import_rows(
        self,
        telegram_id: int,
        rows: Iterable[Iterable[Any]],
        date_format: str,
    ) -> ImportResult:
        result = ImportResult()

        existing = await self._repo.list_by_user(telegram_id)
        records = stays_to_records(existing)

        numbered_rows = iter(enumerate(rows, start=1))
        first_row: tuple[int, list[Any]] | None = None
        for row_number, raw_row in numbered_rows:
            row = _trim_trailing_empty(raw_row)
            if row:
                first_row = (row_number, row)
                break

        if first_row is None:
            result.errors.append(ImportError(row=1, message="file is empty."))
            return result

        first_row_number, first_values = first_row
        columns = _resolve_columns(first_values)
        if columns is None:
            columns = {"country": 0, "entry": 1, "exit": 2}
            import_rows = chain([(first_row_number, first_values)], numbered_rows)
        else:
            import_rows = numbered_rows

        imported_records: list[StayRecord] = []
        as_of = date.today()
        data_row_count = 0

        for i, raw_row in import_rows:
            row = _trim_trailing_empty(raw_row)
            if not row:
                continue

            data_row_count += 1
            if data_row_count > MAX_IMPORT_ROWS:
                result.errors.append(
                    ImportError(
                        row=i,
                        message=f"row limit exceeded (max {MAX_IMPORT_ROWS:,} rows). Remaining rows skipped.",
                    )
                )
                break

            country_input = _cell_text(_cell(row, columns["country"]))
            entry_value = _cell(row, columns["entry"])
            exit_value = _cell(row, columns["exit"])
            exit_str = _cell_text(exit_value)

            # Country validation
            country = resolve_country(country_input)
            if country is None:
                result.errors.append(
                    ImportError(
                        row=i,
                        message=f'unknown country "{html.escape(country_input)}", skipped.',
                    )
                )
                continue

            # Date validation
            entry_date = _parse_date(entry_value, date_format)
            exit_date = _parse_date(exit_value, date_format) if exit_str else None

            if entry_date is None or (exit_str and exit_date is None):
                result.errors.append(
                    ImportError(row=i, message="invalid dates, skipped.")
                )
                continue

            if exit_date is not None and entry_date > exit_date:
                result.errors.append(
                    ImportError(row=i, message="invalid dates, skipped.")
                )
                continue

            # Conflict detection: check against existing stays + already-imported rows
            all_records = records + imported_records

            dup = find_duplicate_entry(all_records, country["code"], entry_date)
            if dup is not None:
                result.errors.append(
                    ImportError(
                        row=i,
                        message=f"overlaps existing {dup.country_name} stay, skipped.",
                    )
                )
                continue

            overlap = find_overlapping_stay(
                all_records, entry_date, exit_date, as_of=as_of
            )
            if overlap is not None:
                result.errors.append(
                    ImportError(
                        row=i,
                        message=f"overlaps existing {overlap.country_name} stay, skipped.",
                    )
                )
                continue

            # Persist stay
            stay = await self._repo.create_entry(
                telegram_id,
                country_code=country["code"],
                country_name=country["name"],
                entry_date=entry_date,
            )

            if exit_date is not None:
                await self._repo.close_stay(stay, exit_date)

            duration = stay_duration_days(entry_date, exit_date, as_of=as_of)
            result.added_entries += 1
            result.countries.add(country["code"])
            result.total_days += duration

            imported_records.append(
                StayRecord(
                    entry_date=entry_date,
                    exit_date=exit_date,
                    country_code=country["code"],
                    country_name=country["name"],
                    stay_id=stay.id,
                )
            )

        return result


def _trim_trailing_empty(row: Iterable[Any]) -> list[Any]:
    values = list(row)
    while values and _cell_text(values[-1]) == "":
        values.pop()
    return values


def _resolve_columns(header: list[Any]) -> dict[str, int] | None:
    normalized = {
        _cell_text(value).casefold(): index for index, value in enumerate(header)
    }
    aliases = {
        "country": ("country",),
        "entry": ("arrival", "date of entry", "entry date"),
        "exit": ("departure", "date of exit", "exit date"),
    }
    columns: dict[str, int] = {}
    for name, candidates in aliases.items():
        index = next((normalized[c] for c in candidates if c in normalized), None)
        if index is None:
            return None
        columns[name] = index
    return columns


def _cell(row: list[Any], index: int) -> Any:
    return row[index] if index < len(row) else None


def _cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _parse_date(value: Any, date_format: str) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(_cell_text(value), date_format).date()
    except (ValueError, TypeError):
        return None
